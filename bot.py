import discord
import datetime
import openpyxl
import asyncio
import random
import string
import os
import youtube_dl
import re
import sys
import bs4
import urllib
import qrcode
import requests
import status
import smtplib
import google_images_download
import xp
import image
import download
import json
import urllib.request
import Dgame
import time
import configparser
import edit
import shutil
import koreanbots
import neispy
import Dtime
import modul
import sqlite3
import psutil
import aiohttp
import mysql.connector
from google_images_download import google_images_download
from email.mime.text import MIMEText
from bs4 import BeautifulSoup
from urllib.request import urlopen
from discord.ext import tasks
from captcha.image import ImageCaptcha
from PIL import Image, ImageDraw, ImageFont
from discord.ext import commands
from discord.utils import get
from urllib.request import Request
from urllib.request import URLError
from urllib.request import HTTPError
from urllib.request import urlopen
from urllib.request import Request, urlopen
from urllib.request import quote
from urllib import parse

client = discord.Client()
Bot = koreanbots.Client(client, 'eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IjY4OTQxOTI2MDg2NjMzMDY0NSIsImlhdCI6MTU5MTUzMzEwMiwiZXhwIjoxNjIzMDkwNzAyfQ.VDrSUhdRuv149oInG9ps1I2p1lVQBRF_Qn71ILx50tqCIMtsn6UhaE_45YM7uxz_ZReCtQ7YxYjzXKDwTskHhSTkygeWGuzkuVwmgny3A3CUYd7iPxRu3EF8EwLTexGS6o5wWAcEVN8kZdDBgQuafMrP3Q588gJNK1THrldwNSI')
queues={}
pre='='
mid=['벤-리스트🚫']
mute=['뮤트/언뮤트']
dsv=['dh-log']
banlist=['꺼져', '뒤져', 'Fuck', 'fuck', 'fucking', 'Fucking', 'What the fuck', 'what the fuck', '씨발', '시바', '시발', '씨바', '좆', '존나', 'Tlqkf', 'tlqkf', 'Tlqk', 'tlqk', '지랄', 'ㅈㄹ', '에미', '애미', '애비', '에비', '조까']
naverapiid="TmlZec6PQAZrzVq_fIao"
naverapipw="fzsgH99SP0"
uptime = Dtime.Client()


admin=445529063528857611
token='Njg5NDE5MjYwODY2MzMwNjQ1.Xxe8Xw.za02Growe1Vq3txtFV93MZ3e8RE'

@client.event
async def bg_change_playing():
    now=datetime.datetime.now()
    webhook = 'https://discordapp.com/api/webhooks/727806784068255755/wJcHemSxcUG0zTn2HSQABbBpDdKExi8xWCe2NLPnZ9HHIFmAJGEZ1q_zJuFGkp_oQOwy'
    requests.post(webhook, {'content':f'메모리:{round(psutil.virtual_memory().used / (1024 * 1024), 2)}MB / {round(psutil.virtual_memory().total / (1024 * 1024), 2)}MB ({psutil.virtual_memory().percent}%)=>{str(now.month)}월{str(now.day)}일{str(now.hour)}시{str(now.minute)}분{str(now.second)}초'})
    await asyncio.sleep(360)
    client.loop.create_task(bg_change_playing())

@client.event
async def on_ready():
    now=datetime.datetime.now()
    print(' ----------------------------------------------------------------------------')
    print(f'| Client : {str(client)}                      |')
    print(f'| Client ID: {client.user.id}                                              |')
    print(f'| Client Name: {client.user.name}                                                        |')
    print(f'| Client Prefix: {pre}                                                           |')
    print(f'| Client Token: {str(token)}  |')
    print(' ----------------------------------------------------------------------------')
    print(f'Clinet StartTimes : {str(now.month)}m {str(now.day)}d {str(now.hour)}h {str(now.minute)}min {str(now.second)}sec')
    now=datetime.datetime.now()
    msg = ['안녕하세요!', f'나는야 {client.user.name}!', f'팀 SB소속 {client.user.name}이에요!\n[스카이봇 포럼](https://discord.gg/57cSTKB)\n[코어 엔터테인먼트](https://discord.gg/TeCpcBq)', f'My name is {client.user.name}!', 'My prefix is /', 'SkyBot은요... 네! 그렇습니다.']
    embed=discord.Embed(
        title='✅ 봇 온라인',
        description=f'|| @here ||\n부팅된 시간 : {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초\n접두사 : {pre}\n\n봇 최초 부팅 핑 : {round(client.latency * 1000)}ms' + f' (0.{round(client.latency * 1000)}seconds)\n\n{random.choice(msg)}\n\n',
        colour=discord.Colour.green()
    )
    embed.set_footer(text=f'{client.user.name} | {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초', icon_url='https://cdn.discordapp.com/attachments/690014483871301647/693625388487671858/8a05d48c12498dea.png')
    await client.get_channel(int(727393651357122562)).send(embed=embed)
    client.loop.create_task(bg_change_playing())

@client.event
async def on_message_delete(message):
    if message.author == client.user:
        return
    elif message.author.bot:
        return
    for channel in message.guild.channels:
        if (str(channel) in dsv):
            try:
                embed=discord.Embed(
                    title=f'메시지 삭제감지',
                    description=message.author.name + " : " + message.content,
                    colour=discord.Colour.red()
                )
                embed.set_footer(icon_url=message.author.avatar_url, text=f'채널 - {message.channel.name}')
                await channel.send(embed=embed)
                if not message.embeds == []:
                    await channel.send(embed=message.embeds[0])
            except:
                return None

@client.event
async def on_message_edit(before, after):
    if before.author.bot:
        return
    if before.content == after.content:
        return
    for channel in before.guild.channels:
        if (str(channel) in dsv):
            try:
                embed=discord.Embed(
                    title=f'메시지 수정감지',
                    description='User : ' + before.author.name + '\n이전 메시지 : ' + before.content + '\n이후 메시지 : ' + after.content,
                    colour=discord.Colour.gold()
                )
                await channel.send(embed=embed)
                if not before.embeds == []:
                    await channel.send(embed=before.embeds[0])
            except:
                return None

@client.event
async def on_member_join(member):
    syschannel = member.guild.system_channel.id
    now=datetime.datetime.now()
    embed=discord.Embed(
        title=f'멤버 입장',
        description=f'{member} ({member.mention})님이 {member.guild}에 들어오셨습니다.\n현재 서버 인원수 : {str(len(member.guild.members))}명',
        colour=discord.Colour.blue()
    ).set_footer(icon_url=member.avatar_url, text=f' | {str(member.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
    embed.set_thumbnail(url=member.avatar_url)
    await client.get_channel(syschannel).send(embed=embed)

@client.event
async def on_member_remove(member):
    syschannel = member.guild.system_channel.id
    now=datetime.datetime.now()
    embed=discord.Embed(
        title=f'멤버 퇴장',
        description=f'{member} ({member.mention})님이 {member.guild}에서 **퇴장** 했습니다.\n현재 서버 인원수 : {str(len(member.guild.members))}명',
        colour=discord.Colour.red()
    ).set_footer(icon_url=member.avatar_url, text=f' | {str(member.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
    embed.set_thumbnail(url=member.avatar_url)
    await client.get_channel(syschannel).send(embed=embed)

@client.event
async def on_voice_channel_join(member):
    for channel in member.guild.channels:
        if (str(channel) in dsv):
            try:
                embed=discord.Embed(
                    title=f'보이스챗 입장 감지',
                    description=f'{member}님이 {member.channel.mention} 채널에 입장했습니다.',
                    colour=discord.Colour.blue()
                ).set_footer(icon_url=member.avatar_url, text=f'채널 - {member.channel.name}')
                await channel.send(embed=embed)
                if not member.embeds == []:
                    await channel.send(embed=member.embeds[0])
            except:
                return None
        elif member.bot:
            return

@client.event
async def on_voice_channel_leave(member):
    for channel in member.guild.channels:
        if (str(channel) in dsv):
            try:
                embed=discord.Embed(
                    title=f'보이스챗 퇴장 감지',
                    description=f'{member}님이 {member.channel.mention} 채널에서 퇴장했습니다.',
                    colour=discord.Colour.blue()
                ).set_footer(icon_url=member.avatar_url, text=f'채널 - {member.channel.name}')
                await channel.send(embed=embed)
                if not member.embeds == []:
                    await channel.send(embed=member.embeds[0])
            except:
                return None
        elif member.bot:
            return

@client.event
async def on_message(message):
    if pre in message.content:
        if message.author.bot:
            return None
        now=datetime.datetime.now()
        embed=discord.Embed(
            title=f'chat in prefix log',
            description=f'User : {message.author} ({message.author.mention})\n\nServer : {message.guild} ({message.guild.id})\n\nChannel : {message.channel.mention} ({message.channel.name}, {message.channel.id})\n\nContent : {message.content}',
            colour=discord.Colour.gold()
        ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
        await client.get_channel(int(727475990946906164)).send(embed=embed)
    else:
        return None
    async def makeembed(title, description):
        now=datetime.datetime.now()
        embed=discord.Embed(
            title=title,
            description=description,
            colour=discord.Colour.green()
        )
        embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
        await message.channel.send(embed=embed)
    for banword in banlist:
        if banword in message.content:
            try:
                await message.delete()
                msg=await message.channel.send(f'{message.author.mention}님은 욕설(비속어)를 사용하여 메세지가 삭제되었습니다.')
                await asyncio.sleep(5)
                await msg.delete()
            except:
                return
            for channel in message.guild.channels:
                if (str(channel) in dsv):
                    try:
                        now=datetime.datetime.now()
                        embed = discord.Embed(
                            title=f'비속어 감지',
                            description=f'메세지 작성인 : {message.author.mention}\n감지된 비속어 : {message.content}\n메세지 감지 채널 : {message.channel.mention}',
                            colour=discord.Colour.red()
                        ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                        await channel.send(embed=embed)
                        if not message.embeds == []:
                            await channel.send(embed=message.embeds[0])
                    except:
                        return None
    try:
        if message.author.bot:
            return None
        
        elif message.content.startswith(f'{pre}따라하기'):
            channel=message.channel_mentions[0]
            text=message.content[28:]
            author=message.author.name
            embed=discord.Embed(
                title=f'**{author}**',
                description=f'{text}',
                colour=discord.Colour.blue()
            )
            await message.author.send(f"{channel} 채널에 {text} 메시지를 전송했습니다.")
            await channel.send(embed=embed)

        elif message.content.startswith(f'{pre}이미지'):
                Text = ""
                learn = message.content.split(" ")
                vrsize = len(learn)  # 배열크기
                vrsize = int(vrsize)
                for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
                    Text = Text + " " + learn[i]
                print(Text.strip())  # 입력한 명령어

                randomNum = random.randrange(0, 3) # 랜덤 이미지 숫자

                location = Text
                enc_location = urllib.parse.quote(location) # 한글을 url에 사용하게끔 형식을 바꿔줍니다. 그냥 한글로 쓰면 실행이 안됩니다.
                hdr = {'User-Agent': 'Mozilla/5.0'}
                # 크롤링 하는데 있어서 가끔씩 안되는 사이트가 있습니다.
                # 그 이유는 사이트가 접속하는 상대를 봇으로 인식하였기 때문인데
                # 이 코드는 자신이 봇이 아닌것을 증명하여 사이트에 접속이 가능해집니다!
                url = 'https://search.naver.com/search.naver?where=image&sm=tab_jum&query=' + enc_location # 이미지 검색링크+검색할 키워드
                req = Request(url, headers=hdr)
                html = urllib.request.urlopen(req)
                bsObj = bs4.BeautifulSoup(html, "html.parser") # 전체 html 코드를 가져옵니다.
                # print(bsObj)
                imgfind1 = bsObj.find('div', {'class': 'photo_grid _box'}) # bsjObj에서 div class : photo_grid_box 의 코드를 가져옵니다.
                # print(imgfind1)
                imgfind2 = imgfind1.findAll('a', {'class': 'thumb _thumb'}) # imgfind1 에서 모든 a태그 코드를 가져옵니다.
                imgfind3 = imgfind2[randomNum]  # 0이면 1번째사진 1이면 2번째사진 형식으로 하나의 사진 코드만 가져옵니다.
                imgfind4 = imgfind3.find('img') # imgfind3 에서 img코드만 가져옵니다.
                imgsrc = imgfind4.get('data-source') # imgfind4 에서 data-source(사진링크) 의 값만 가져옵니다.
                embed = discord.Embed(
                    title=f"{Text}이미지 검색 결과입니다!", 
                    timestamp=message.created_at, 
                    colour=discord.Colour.green()
                )
                embed.set_image(url=imgsrc) # 이미지의 링크를 지정해 이미지를 설정합니다.
                await message.channel.send(embed=embed) # 메시지를 보냅니다.


        elif message.content == f'{pre}dh-log':
            await message.guild.create_text_channel(name='dh-log')
            await makeembed('생성 완료!', '`dh-log` 채널을 생성했습니다!')

        elif message.content == f'{pre}슬로우모드':
            sec = message.content[7:]
            if sec == None:
                await makeembed('오류!', '초를 적어주십시오.')
                return None
            elif sec == ' ':
                await makeembed('오류!', '공백.')
                return None
            elif sec < int(61):
                await makeembed('오류!', '테러방지를 위한 60초 까지.')
                return None
            else:
                await makeembed('오류!', 'Unknow')
                return None
            await message.channel.edit(slowmode_delay=int(sec))
            await makeembed('슬로우모드 활성화.', f'현재 채널은 슬로우 모드가 걸렸습니다.\n딜레이 : **{sec}**')
            if sec == '취소':
                await message.channel.edit(slowmode_delay=int(0))
                await makeembed('취소!', '슬로우모드를 0초 로 변경했습니다.')

        elif message.content.startswith(f"{pre}채널정보"):
            if len(message.channel_mentions) == 0:
                channel = message.channel
            else:
                channel = message.channel_mentions[0]
            name = channel.name
            cid = channel.id
            topic = channel.topic # ctpye 종류 pos 순서 topic 주제 cid 채널아이디
            if topic == "" or topic == None:
                topic = "없음"
            pos = str(channel.position+1) + "번"
            ctype = str(channel.type)
            created = str(channel.created_at)
            embed=discord.Embed(
                title=f'{name} 채널 정보',
                description=f'{name.mention}\n아이디 : {cid}\n순서 : {pos}\n주제 : {topic}\n종류 : {ctype}\n생성일 : {created}',
                colour=discord.Colour.blue()
            ).set_footer(icon_url=message.author.avatar_url, text=f'| {message.author.display_name} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
            await message.channel.send(embed=embed)

        elif message.content == f'{pre}시스템':
            await makeembed(f'메모리:{round(psutil.virtual_memory().used / (1024 * 1024), 2)}MB / {round(psutil.virtual_memory().total / (1024 * 1024), 2)}MB ({psutil.virtual_memory().percent}%)', f'=>{str(now.month)}월{str(now.day)}일{str(now.hour)}시{str(now.minute)}분{str(now.second)}초')
        
        elif message.content == f'{pre}이모지':
            msg = message.content[5:].split(' ' and '<' and ':' and '>')
            emoji = message.guild.get_emoji(msg)
            if emoji not in '<' or ':' or '>':
                await makeembed('이모지가 아닙니다!', '이모지가 아닙니다. 다시 시도하세요.')
                return None
            try:
                await message.channel.send(f'https://cdn.discordapp.com/emojis/{emoji}.png?v=1')
            except:
                await makeembed('오류', '오류')

        elif message.content.startswith(f"{pre}계산"):
            channel = message.channel
            math = message.content[4:]
            if math == "":
                await message.channel.send('계산식를 입력해주세요')
            elif len(message.mentions) >= 1 or len(message.role_mentions) >= 1 or len(message.channel_mentions) >= 1:
                await message.channel.send('계산식이 올바르지 않습니다..')
            else:
                mathtext = ""
                allowed = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0", ".", "+", "-", "*", "/", "(", ")"]
                for i in math:
                    if i in allowed:
                        mathtext += i
                    else:
                        mathtext += ""
                try:
                    value = eval(mathtext)
                    embed=discord.Embed(
                        title=f'{mathtext}식의 결과',
                        description=f'{str(value)}',
                        colour=discord.Colour.blue()
                    )
                    await message.channel.send(embed=embed)
                except:
                    await message.channel.send('계산식이 올바르지 않습니다.')

        elif message.content.startswith(f'{pre}실검'):
            json=requests.get("https://www.naver.com/srchrank?frm=main").json()
            embed = discord.Embed(
                title='네이버 실시간 검색어',
                description='현재 시각 기준 네이버의 실시간 검색어입니다.',
                colour=discord.Colour.green()
            )
            for r in json.get('data'):
                a=r.get("keyword")
                embed.add_field(name=f'{r.get("rank")}위', value=r.get("keyword"), inline=False)
            embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} |')
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}타이머"):
            channel = message.channel
            if len(message.content[5:].split()) == 0:
                await message.channel.send('1~2개의 인수가 필요함')
            else:
                time = message.content[5:].split()[0]
                if len(message.content[5:].split()) >= 2:
                    reason = message.content[5:].split()[1]
                else:
                    reason = "사유 없음"
                try:
                    time = int(time)
                    embed=discord.Embed(
                        title=f'타이머',
                        description=f':alarm_clock: {reason} | {str(time)}초(이)라는 타이머가 등록되었습니다.',
                        colour=0x00f000
                    )
                    embed.set_footer(icon_url=message.author.avatar_url, text=f'{message.author}에게 요청받음')
                    await message.channel.send(embed=embed)
                    await asyncio.sleep(time)
                    try:
                        await message.channel.send(f'{message.author.mention}')
                        embed=discord.Embed(
                            title=f'타이머 (끝)',
                            description=f':alarm_clock: {reason}',
                            colour=0x00f000
                        )
                        embed.set_footer(icon_url=message.author.avatar_url, text=f'{message.author}에게 요청받음')
                        await message.channel.send(embed=embed)
                    except:
                        embed=discord.Embed(
                            title=f'타이머 (끝)',
                            description=f':alarm_clock: {reason}',
                            colour=0x00f000
                        )
                        embed.set_footer(icon_url=message.author.avatar_url, text=f'{message.author}에게 요청받음')
                        await message.channel.send(embed=embed)
                except:
                    await message.channel.send('계산 명령어의 시간(초)는 숫자여야만 합니다!')

        elif message.content == f'{pre}업타임':
            await makeembed(f"{uptime.hours()}시간 {uptime.minitues()}분 {uptime.seconds()}초", '다해봇 업타임 기능')

        elif message.content.startswith(f'{pre}랜덤'):
            msg = message.content[4:].split("/")
            await makeembed('결과', f'{str(random.choice(msg))}')
        
        elif message.content.startswith(f'{pre}단축'):
            target=message.content[4:]
            client_id="TmlZec6PQAZrzVq_fIao"
            client_secret="fzsgH99SP0"
            header = {'X-Naver-Client-Id': client_id, 'X-Naver-Client-Secret': client_secret}
            naver = 'https://openapi.naver.com/v1/util/shorturl'
            data = {'url': target}
            maker=requests.post(url=naver,data=data,headers=header)
            maker.close()
            output=maker.json()['result']['url']
            await makeembed('링크 단축', f'{output}')

        elif message.content == f'{pre}코로나':
            url = 'https://coronamap.site/'
            res = requests.get(url)
            soup = BeautifulSoup(res.text, 'html.parser')
            hwagjin = soup.find('div', class_ = 'content').find('div').text
            dead = str(soup.find('div', class_ = 'content1 clear')).split('</div>')[5].split('<div>')[1]
            noknow = BeautifulSoup(requests.get('https://search.naver.com/search.naver?sm=top_hty&fbm=0&ie=utf8&query=%EC%BD%94%EB%A1%9C%EB%82%98').text, 'html.parser').find('div', class_ = 'graph_view').find('div', class_ = 'circle orange level5').find('strong', class_ = 'num').text
            none = str(soup.find('div', class_ = 'content1 clear')).split('</div>')[1].split('<div>')[1]
            embed = discord.Embed(
                title = '코로나19 현황',
                color = discord.Color.blue(),
                description = f'⛔️확진자: {hwagjin}\n⭕️검사진행: {noknow}\n🩸사망자: {dead}\n🆓격리해제: {none}'
            )
            await message.channel.send(embed = embed)

        elif message.content.startswith(f"{pre}역사"):
            now=datetime.datetime.now()
            msg=await makeembed('This Bot`s history', 'name : ChatBot -> 다해봇> 2020. 04. 06일 변환')
            msg1=await makeembed('this Bot`s birthday', '2020. 03. 17일')
            await msg.delete()
            await msg1.delete()

        elif message.content == f'{pre}새총':
            now=datetime.datetime.now()
            fly = ['UN10', '선물', '선물', '선물', '비행기', '비행기', '까마귀', '까마귀', '까마귀', '까마귀', '까치', '까치', '까치', '까치', '까치', '참새', '참새', '참새', '참새', '참새', '참새', '사람', '사람', '사람', '사람', '사람', '사람', '자동차', '자동차', '자동차', '자동차', '자동차', '자동차', '자동차', '자동차']
            flynumber = random.choice(fly)
            await makeembed('새총 발사중...', '휘용요요요요용')
            await asyncio.sleep(5)
            await makeembed('탁!', f'{message.author} : 뭐지?')
            await asyncio.sleep(2)
            if flynumber == 'UN10':
                await message.channel.send(f'{message.author.mention}')
                await makeembed(f'엇! 야생의 UN10이 나타났다!', f'아닛! {client.user.name}이 떨어틀인거 같네...')
                await client.get_user(int(445529063528857611)).send(f'{message.author} 가 니 뽑음 ㅅㄱ;; ㅋ')
            if flynumber == '선물':
                await message.channel.send(f'{message.author.mention}')
                await makeembed('응? 선물이네...', '산타가 지나갔나?')
            if flynumber == '비행기':
                await message.channel.send(f'{message.author.mention}')
                await makeembed('헐...', '(멀리서)삐용삐용삐용')
            if flynumber == '까마귀':
                embed = discord.Embed(
                    title=f'까마귀',
                    description=f'새를 잡았다!',
                    colour=discord.Colour.red()
                )
                embed.add_field(name=f'크기', value=f'30.5cm', inline=True)
                embed.add_field(name=f'희귀도', value=f'희귀')
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            if flynumber == '까치':
                embed = discord.Embed(
                    title=f'까치',
                    description=f'새를 잡았다!',
                    colour=discord.Colour.green()
                )
                embed.add_field(name=f'크기', value=f'25.0cm', inline=True)
                embed.add_field(name=f'희귀도', value=f'보통')
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            if flynumber == '참새':
                embed = discord.Embed(
                    title=f'참새',
                    description=f'새를 잡았다!',
                    color=0xFFFFFF
                )
                embed.add_field(name=f'크기', value=f'10.5cm', inline=True)
                embed.add_field(name=f'희귀도', value=f'일반')
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            if flynumber == '사람':
                await message.channel.send(f'{message.author.mention}')
                await message.channel.send(f'삐용삐용\n{message.author} : 헉... 큰일났다! 튀어!')
            if flynumber == '자동차':
                await message.channel.send(f'{message.author.mention}')
                await message.channel.send(f'쾅\n{message.author} : 안돼!! 1억3000만원 하는 UBN10 자통차잖아!')

        elif message.content == f'{pre}주사위':
            do = [1, 2, 3, 4, 5, 6]
            num_elements = do
            await makeembed('주사위를 던집니다!', '굴리는중...')
            await asyncio.sleep(3)
            await makeembed('주사위 결과 : ', f'결과 : {random.choice(do)}, \n\n주사위에 있는 숫자 : {num_elements}')

        elif message.content.startswith(f"{pre}정보"):
            now=datetime.datetime.now()
            author=message.mentions[0]
            #0x00ff00
            roles = [role for role in author.roles]
            date = datetime.datetime.utcfromtimestamp(((int(author.id) >> 22) + 1420070400000) / 1000)
            status_dict: dict = {discord.Status.online: '✅ 온라인',
                discord.Status.offline: '⚫ 오프라인',
                discord.Status.idle: "🔑 자리비움",
                discord.Status.do_not_disturb: "⛔ 방해금지"}
            user_status = status_dict[author.status]
            embed = discord.Embed(colour=discord.Colour.green(), title=f'{author} 님의 정보')
            embed.add_field(name='이름', value=f'{author}', inline=False) #inline True : 옆
            embed.add_field(name='서버닉네임', value=author.display_name, inline=False)
            embed.add_field(name='아이디', value=f'{author.id}')
            embed.add_field(name='가입일', value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일" + str(date.hour) + "시" + str(date.minute) + "분" + str(date.second) + "초", inline=True)
            embed.add_field(name='서버 가입일', value=f'{author.joined_at.year}년 {author.joined_at.month}월 {author.joined_at.day}일 {author.joined_at.hour}시 {author.joined_at.minute}분 {author.joined_at.second}초')
            embed.add_field(name='아이디', value=author.id, inline=False)
            embed.add_field(name=f"가진 역할들({len(roles)-1}개)", value=f" ".join([role.mention for role in roles][1:]), inline=False)
            embed.add_field(name="가장 높은 역할", value=f"{author.top_role.mention}", inline=False)
            embed.add_field(name='현재 상태', value=f'{user_status}', inline=False)
            if author.bot == True:
                embed.add_field(name=f'봇', value=f'True')
                embed.add_field(name='초대 코드', value='https://discord.com/oauth2/authorize?client_id=' + str(author.id) + '&scope=bot&permissions=2147483639')
            else:
                embed.add_field(name=f'봇', value=f'False')
            embed.add_field(name='하는 것', value=author.activity, inline=False)
            embed.set_thumbnail(url=author.avatar_url)
            embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
            await message.channel.send(embed=embed)
            if author.guild_permissions.administrator: #관리자
                per1 = '관리자 : O'
            else:
                per1 = '관리자 : X'
            
            if author.guild_permissions.view_audit_log: #감사로그 보기
                per2 = '감사로그 보기 : O'
            else:
                per2 = '감사로그 보기 : X'

            if author.guild_permissions.manage_guild: #서버 관리하기
                per3 = '서버 관리하기 : O'
            else:
                per3 = '서버 관리하기 : X'

            if author.guild_permissions.manage_roles: #역할 관리하기
                per4 = '역할 관리하기 : O'
            else:
                per4 = '역할 관리하기 : X'

            if author.guild_permissions.manage_channels: #채널 관리하기
                per5 = '채널 관리하기 : O'
            else:
                per5 = '채널 관리하기 : X'

            if author.guild_permissions.kick_members: #멤버 추방하기
                per6 = '멤버 추방하기 : O'
            else:
                per6 = '멤버 차단하기 : X'

            if author.guild_permissions.ban_members: #멤버 차단하기
                per7 = '멤버 차단하기 : O'
            else:
                per7 = '멤버 차단하기 : X'

            if author.guild_permissions.create_instant_invite: #채널 관리하기
                per8 = '초대 코드 만들기 : O'
            else:
                per8 = '초대 코드 만들기 : X'

            if author.guild_permissions.change_nickname: #별명 변경하기
                per9 = '별명 변경하기 : O'
            else:
                per9 = '별명 변경하기 : X'

            if author.guild_permissions.manage_nicknames: #별명 관리하기
                per10 = '별명 관리하기 : O'
            else:
                per10 = '별명 관리하기 : X'

            if author.guild_permissions.manage_emojis: #이모티콘 관리
                per11 = '이모티콘 관리 : O'
            else:
                per11 = '이모티콘 관리 : X'

            if author.guild_permissions.view_channel: #채팅 채널 읽기 및 음성 채널 보기
                per12 = '채팅 채널 읽기 및 음성 채널 보기 : O'
            else:
                per12 = '채팅 채널 읽기 및 음성 채널 보기 : X'

            if author.guild_permissions.send_messages: #메세지 보내기
                per13 = '메세지 보내기 : O'
            else:
                per13 = '메세지 보내기 : X'

            if author.guild_permissions.send_tts_messages: #tts메세지 보내기
                per14 = 'TTS 메세지 보내기 : O'
            else:
                per14 = 'TTS 메세지 보내기 : X'

            if author.guild_permissions.manage_messages: #메세지 관리하기
                per15 = '메세지 관리하기 : O'
            else:
                per15 = '메세지 관리하기 : X'

            if author.guild_permissions.embed_links: #링크 첨부하기
                per16 = '링크 첨부하기 : O'
            else:
                per16 = '링크 첨부하기 : X'

            if author.guild_permissions.attach_files: #파일 첨부하기
                per17 = '파일 첨부하기 : O'
            else:
                per17 = '파일 첨부하기 : X'

            if author.guild_permissions.read_message_history: #메세지 기록보기
                per18 = '메세지 기록보기 : O'
            else:
                per18 = '메세지 기록보기 : X'

            if author.guild_permissions.mention_everyone: #@everyone, @here 모든 역할 멘션하기
                per19 = '@everyone, @here 모든 역할 멘션하기 : O'
            else:
                per19 = '@everyone, @here 모든 역할 멘션하기 : X'

            if author.guild_permissions.use_external_emojis: #외부 이모티콘 사용하기
                per20 = '외부 이모티콘 사용하기 : O'
            else:
                per20 = '외부 이모티콘 사용하기 : X'

            if author.guild_permissions.add_reactions: #반응 추가하기
                per21 = '반응 추가하기 : O'
            else:
                per21 = '반응 추가하기 : X'

            if author.guild_permissions.connect: #연결
                per22 = '연결 - 음성 : O'
            else:
                per22 = '연결 - 음성 : X'

            if author.guild_permissions.speak: #말하기
                per23 = '말하기 - 음성 : O'
            else:
                per23 = '말하기 - 음성: X'

            if author.guild_permissions.stream: #동영상
                per24 = '동영상 - Go Live : O'
            else:
                per24 = '동영상 - Go Live : X'

            if author.guild_permissions.mute_members: #멤버들의 마이크 음소거하기
                per22 = '멤버들의 마이크 음소거하기 - 음성 : O'
            else:
                per22 = '멤버들의 마이크 음소거하기 - 음성 : X'

            if author.guild_permissions.deafen_members: #멤버의 헤드셋 음소거하기
                per22 = '멤버의 헤드셋 음소거하기 - 음성 : O'
            else:
                per22 = '멤버의 헤드셋 음소거하기 - 음성 : X'

            if author.guild_permissions.move_members: #멤버 이동
                per23 = '멤버 이동 - 음성 : O'
            else:
                per23 = '멤버 이동 - 음성: X'

            if message.author.guild_permissions.priority_speaker: #말하기
                per23 = '우선 발언권 - 음성 : O'
            else:
                per23 = '우선 발언권 - 음성: X'

            embed=discord.Embed(
                title=f'{author}님의 {message.guild}에서의 권한',
                description=f'{per1}\n{per2}\n{per3}\n{per4}\n{per5}\n{per6}\n{per7}\n{per8}\n{per9}\n{per10}\n{per11}\n{per12}\n{per13}\n{per14}\n{per15}\n{per16}\n{per17}\n{per18}\n{per19}\n{per20}\n{per21}\n{per22}\n{per23}\n{per24}\n',
                colour=discord.Colour.blue()
            ).set_footer(icon_url=author.avatar_url, text=f'| {author.display_name} | {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
            await message.channel.send(embed=embed)

        if message.content.startswith("=컴파일"):
            if message.author.id == 445529063528857611:
                a=message.content[5:]
                try:
                    msg=await message.channel.send(embed=discord.Embed(color=0x2F3136, title="머리를 깍는 중...",description=f"""📥INPUT📥```py\n{a}```📤OUTPUT📤```pyevaling...```"""))
                    aa=await eval(a)
                except Exception as e:
                    await msg.edit(embed=discord.Embed(color=0x2F3136, title="결과",description=f"""📥INPUT📥```py\n{a}```📤OUTPUT📤```py\n{e}```"""))
                    try:
                        aa = eval(a)
                    except Exception as e:
                        await msg.edit(embed=discord.Embed(color=0x2F3136, title="결과",description=f"""📥INPUT📥```py\n{a}```📤OUTPUT📤```py\n{e}```"""))
                    else:
                        await msg.edit(embed=discord.Embed(color=0x2F3136, title=f"결과",description=f"""📥INPUT📥```py\n{a}```📤OUTPUT📤```py\n{aa}```""")) 
                else:
                    await msg.edit(embed=discord.Embed(color=0x2F3136, title="결과",description=f"""📥INPUT📥```py\n{a}```📤OUTPUT📤```py\n{aa}```"""))
            else:
                await message.channel.send("권한없음")

        elif message.content.startswith(f"{pre}서버정보"):
            now=datetime.datetime.now()
            roles = [role for role in message.guild.roles]
            embed = discord.Embed(color=0x00D8FF, title=f'{message.guild}의 정보')
            embed.add_field(name='서버 주인', value=f'{message.guild.owner} ({message.guild.owner_id})', inline=False)
            embed.add_field(name='서버 생성일', value=f'{message.guild.created_at.year}년 {message.guild.created_at.month}월 {message.guild.created_at.day}일 {message.guild.created_at.hour}시 {message.guild.created_at.minute}분 {message.guild.created_at.second}초', inline=False)
            embed.add_field(name='서버인원', value=f'{str(len(message.guild.members))}명', inline=False)
            embed.add_field(name='서버지역', value=f'{message.guild.region}')
            embed.add_field(name='서버 아이디', value=f'{message.guild.id}', inline=False)
            embed.add_field(name="서버 부스트 레벨", value=f'{message.guild.premium_tier}레벨', inline=False)
            embed.add_field(name="서버 부스트 횟수", value=f' {message.guild.premium_subscription_count}번', inline=False)
            if message.guild.afk_channel != None:
                embed.add_field(name = f'🏴잠수 채널', value = f'잠수 채널이 있습니다.\n{message.guild.afk_channel.name} (타이머: {message.guild.afk_timeout}초)', inline = False)
            else:
                embed.add_field(name="🏴잠수 채널", value=":잠수 채널이 존재하지 않습니다.", inline=False)
            if message.guild.system_channel != None:
                embed.add_field(name = f'🌌시스템 채널', value = f'시스템 채널이 있습니다.\n<#{message.guild.system_channel.id}>', inline = False)
            else:
                embed.add_field(name="🌌시스템 채널", value="시스템 채널이 존재하지 않습니다!", inline=False)
            embed.add_field(name="서버 채널 수", value=f'전체 채널: {len(message.guild.channels)}개 \n(채팅채널 : {len(message.guild.text_channels)}개 | 음성채널 : {len(message.guild.voice_channels)}개 | 카테고리 : {len(message.guild.categories)}개)', inline=False)
            await message.channel.send(embed=embed)
            result = ""
            for emoji in message.guild.emojis:
                result += str(emoji)
            embed = discord.Embed(color=0x00D8FF, title=f'서버이모지')
            embed.add_field(name='커스텀 이모지', value=f'({str(result)})')
            embed.add_field(name='갯수', value=f'{str(len(message.guild.emojis))}개')
            await message.channel.send(embed=embed)
            embed = discord.Embed(color=0x00D8FF, title='서버역할')
            embed.add_field(name=f"역할({len(roles)-1}개)", value=f" ".join([role.mention for role in roles][1:]), inline=False)
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}프사"):
            if not message.mentions:
                await makeembed('어케 알려줌', '?')
                return
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author.id == 445529063528857611:
                now=datetime.datetime.now()
                author = message.mentions[0]
                embed = discord.Embed(color=0x00ff00, title=f'{author} 님의 프로필 사진')
                embed.add_field(name=f'{author} 님의 프로필 사진', value=f'{author}', inline=False)
                embed.set_image(url=author.avatar_url)
                embed.set_footer(text=f'{client.user.name} | {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초', icon_url=message.author.avatar_url)
                await message.channel.send(embed=embed)
            else:
                await makeembed(f'님은 사용불가', '사진을 맘대로 사용할거 같아서 ADMIN만 사용할수 있게 했어요!')
                return None

        elif message.content.startswith(f"{pre}도움말"):
            await makeembed(f'{pre}명령어', '이 명령어로 모든 명령어를 알아보세요!')

        elif message.content.startswith(f'{pre}가위바위보'):
            msg = message.content[7:]
            rac = ['가위', '바위', '보']
            racran = random.choice(rac)
            if msg == '가위':
                if racran == '가위':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n비김'
                    )
                    await message.channel.send(embed=embed)
                if racran == '바위':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n졌음',
                        colour=discord.Colour.red()
                    )
                    await message.channel.send(embed=embed)
                if racran == '보':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n이김',
                        colour=discord.Colour.gold()
                    )
                    await message.channel.send(embed=embed)
            if msg == '바위':
                if racran == '가위':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n이김',
                        colour=discord.Colour.gold()
                    )
                    await message.channel.send(embed=embed)
                if racran == '바위':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n비김'
                    )
                    await message.channel.send(embed=embed)
                if racran == '보':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n졌음',
                        colour=discord.Colour.red()
                    )
                    await message.channel.send(embed=embed)
            if msg == '보':
                if racran == '가위':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n졌음',
                        colour=discord.Colour.red()
                    )
                    await message.channel.send(embed=embed)
                if racran == '바위':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n이김',
                        colour=discord.Colour.gold()
                    )
                    await message.channel.send(embed=embed)
                if racran == '보':
                    embed=discord.Embed(
                        title=f'가위바위보 기능',
                        description=f'**{message.author}** :\n{msg}\n **{client.user.name}** :\n{racran} \n\n비김',
                    )
                    await message.channel.send(embed=embed)

        elif message.content.startswith(f'{pre}오늘뭐먹지'):
            food = "베이컨 감자 김치찌개 라면 쌀 계란 나물 시금치 계란과자 보리차 물 굶어 햄버거 치킨 피자 탕수육 짜장면 초밥 삼겹살 소고기 감자튀김 콜라 사이다 두부 계란밥"
            foodchoice = food.split(' ')
            foodnumber = random.randint(1, len(foodchoice))
            foodresult = foodchoice[foodnumber-1]
            await makeembed('오늘 당신이 먹을것은?', f'{foodresult}\n입니다.')

        elif message.content.startswith(f'{pre}배워'):
            file = openpyxl.load_workbook("기억.xlsx")
            sheet = file.active
            learn = message.content.split(" ")
            for i in range(1, 51):
                if sheet["A" + str(i)].value == "-":
                    sheet["A" + str(i)].value = learn[1]
                    sheet["B" + str(i)].value = learn[2]
                    await channel.send("단어 학습완료")
                    break
            file.save('기억.xlsx')

        elif message.content.startswith(f'{pre}기억'):
            file = openpyxl.load_workbook("기억.xlsx")
            sheet = file.active
            memory = message.content.split(' ')
            for i in range(1, 51):
                if sheet["A" + str(i)].value == memory[1]:
                    await message.channel.send(sheet["B" + str(i)].value)
                    break

        elif message.content.startswith(f"{pre}투표"):
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author.id == 445529063528857611:
                await message.delete()
                vote = message.content[4:].split("/")
                if vote == '':
                    await makeembed('내용을 입력하세요!', '투표기능')
                await message.channel.send("📣투표 - " + vote[0])
                for i in range(1, len(vote)):
                    choose = await message.channel.send("```" + vote[i] + "```")
                    await choose.add_reaction('✔️')
            else:
                await makeembed('사용 ㄴㄴ', '관리자 기능')
                return None

        elif message.content.startswith(f'{pre}채널메시지'):
            if not message.channel_mentions:
                await makeembed('보낼 채널을 언급해주세요', '관리자 기능')
                return
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author. id == 445529063528857611:
                channel = message.channel_mentions[0]
                msg = message.content[29:]
                if msg == '':
                    await makeembed('메시지를 입력하세요', '관리자 기능')
                    return
                await channel.send(msg)
            else:
                await makeembed('사용불가', '관리자 기능')
                return None

        elif message.content.startswith(f"{pre}채널만들기"):
            if message.author.guild_permissions.administrator or message.author.id == 445529063528857611:
                msg = message.content[7:]
                await message.guild.create_text_channel(name=msg)
                await makeembed(f'{msg}', '생성완료!')
            else:
                await makeembed(f'{message.author.mention}', '님은 사용불가~')
                return None

        elif message.content.startswith(f"{pre}보이스챗"):
            if message.author.guild_permissions.administrator or message.author.id == 445529063528857611:
                msg = message.content[5:]
                await message.guild.create_voice_channel(name=msg)
                await makeembed(f'{msg}', '생성완료!')
            else:
                await makeembed(f'{message.author.mention}', '님은 사용불가~')
                return None

        elif message.content.startswith(f"{pre}ban"):
            if not message.mentions:
                await makeembed('유저를 맨션해주세요!!', 'Error')
                return
            if message.author.guild_permissions.administrator or message.author.guild_permissions.ban_members or message.author.id == 445529063528857611:
                now = datetime.datetime.now()
                author = message.mentions[0]
                msg = message.content[29:]
                case = 0
                case += 1
                embed = discord.Embed(color=0xFF0000, title=f'You are Ban from this server({message.guild})!')
                embed.add_field(name='사유 : ', value=msg, inline=False)
                embed.add_field(name='억울하시다면...', value=f'{message.guild} 서버 관리자에게 문의 하세요!', inline=False)
                embed.set_footer(icon_url=author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                await author.send(embed=embed)
                for channel in message.guild.channels:
                    if (str(channel) in mid):
                        try:
                            embed = discord.Embed(
                                title=f'{author} is ban from {message.guild} server!',
                                description=f'User : {author} ({author.mention})\nAdmin : {message.author}\nReason : ' + msg,
                                color=discord.Colour.red()
                            )
                            embed.set_footer(icon_url=author.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                            await message.guild.ban(author)
                            await message.channel.send(embed=embed)
                            if not message.embeds == []:
                                await channel.send(embed=message.embeds[0])
                        except:
                            return None
            else:
                await makeembed(f'{message.author.mention}님은...', '이 명령어를 쓰기엔 하찮아요...')

        elif message.content.startswith(f"{pre}개별공지"):
            if message.author.guild_permissions.administrator or message.author.id == 445529063528857611:
                msg = message.content[6:]
                now=datetime.datetime.now()
                embed=discord.Embed(
                    title=f'{message.guild}서버의 개별공지',
                    description=f'Admin : {message.author} ({message.author.mention})\n\nContents :\n{msg}',
                    colour=discord.Colour.green()
                )
                embed.set_footer(icon_url=author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                await author.send(embed=embed)
                for u in message.guild.users:
                    await u.send(embed = embed)
            else:
                await makeembed('관리자 권한이 없습니다!', f'{message.author.mention}(((퍽')

        elif message.content.startswith(f'{pre}DM공지'):
            author = message.mentions[0]
            msg = message.content[29:]
            now=datetime.datetime.now()
            if message.author.guild_permissions.administrator or message.author.id == 445529063528857611:
                embed=discord.Embed(
                    title=f'{message.guild}서버의 DM공지',
                    description=f'Admin : {message.author} ({message.author.mention})\n\nUser : {author} ({author.mention})\n\nContents :\n\n{msg}',
                    colour=discord.Colour.blue()
                )
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                await author.send(embed=embed)
                await makeembed('전송완료!', '메시지 전송을 완료했습니다.')

        elif message.content.startswith(f"{pre}스킨"):
            nickname = message.content[4:]
            channel = message.channel
            embed = discord.Embed(
                title = f'{nickname}님의 스킨',
                description = f'[[ 아바타 ]](https://minotar.net/helm/{nickname}/100.png) [[ 큐브 아바타 ]](https://minotar.net/cube/{nickname}/100.png) \n[[ 전신 ]](https://minotar.net/armor/body/{nickname}/100.png) [[ 반신 ]](https://minotar.net/armor/bust/{nickname}/100.png)\n[[ 스킨 다운로드 ]](https://minotar.net/download/{nickname})',
                colour = discord.Colour.green()
            ).set_thumbnail(url = f"https://minotar.net/armor/bust/{nickname}/100.png")
            await message.channel.send(embed = embed)

        elif message.content.startswith(f"{pre}kick"):
            if not message.mentions:
                await makeembed('유저를 맨션해주세요!', 'Error')
                return
            if message.author.guild_permissions.administrator or message.author.guild_permissions.kick_members or message.author.id == 445529063528867611:
                now = datetime.datetime.now()
                author = message.mentions[0]
                msg = message.content[29:]
                for channel in message.guild.channels:
                    if (str(channel) in mid):
                        try:
                            embed = discord.Embed(
                                title=f'{author} is kick from {message.guild} server!',
                                description=f'User : {author} ({author.mention})\nAdmin : {message.author}\nReason : ' + msg,
                                color=discord.Colour.gold()
                            )
                            embed.set_footer(icon_url=author.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                            await message.guild.kick(author)
                            await message.channel.send(embed=embed)
                            if not message.embeds == []:
                                await channel.send(embed=message.embeds[0])
                        except:
                            return None
                embed = discord.Embed(color=0xFF0000, title=f'You are kick from this server!({message.guild})')
                embed.add_field(name='사유 : ', value=msg)
                embed.add_field(name='억울하시다면...', value='서버 관리자에게 문의 하세요!')
                await author.send(embed=embed)

        elif message.content.startswith(f"{pre}역할"):
            author = message.mention[0]
            msg = message.content[4:]
            role = discord.utils.get(message.guild.roles, name=msg) #await author.add_role(role)
            if not message.mentions:
                await makeembed('유저를 멘션해주세요!', f'{pre}역할 <역할이름> <멘션>')
                return
            if msg == None:
                await makeembed('역할이름을 적어주세요!', f'{pre}역할 <역할이름> <멘션>')
            if message.author.guild_permissions.administrator or message.author.id == 445529063528857611:
                for channel in message.guild.channels:
                    if (str(channel) in mid):
                        try:
                            embed=discord.Embed(
                                title=f'역할 추가 감지',
                                description=f'User : {author}({author.mention})\nAdmin : {message.author}\nRole Name : {msg}',
                                colour=discord.Colour.blue()
                            )
                            await channel.send(embed=embed)
                            if not message.embeds == []:
                                await channel.send(embed=message.embeds[0])
                        except:
                            return None
            else:
                await makeembed('니 권한 없음 ㅅㄱ', 'ㅋ •••...')
                return None

        elif message.content.startswith(f"{pre}DM"):
            author = message.mentions[0]
            msg = message.content[26:]
            if msg == '':
                await makeembed('내용을 적어야 보내든 말든하지', 'ㅋ')
                return
            if not message.mentions:
                await makeembed('누구한테', '보낼겨')
                return
            embed = discord.Embed(color=0x00ff00, title=f'{author}')
            embed.add_field(name='보낸이 : ', value=message.author.name, inline=False)
            embed.add_field(name='내용 :', value=msg, inline=True)
            embed.set_thumbnail(url=message.author.avatar_url)
            embed.set_footer(text=f"{message.author}", icon_url=message.author.avatar_url)
            choose=await message.channel.send('메시지 보낼거임? O or X')
            await choose.add_reaction("⭕")
            await choose.add_reaction("❌")
            def c(reaction, user):
                return user == message.author and str(reaction) in ["⭕", "❌"] and choose.id == reaction.message.id # 이모지 리액션 부분
            reaction, user = await client.wait_for('reaction_add',check=c)# 이모지 리액션 부분
            if str(reaction.emoji) == '⭕':
                await choose.delete()
                await author.send(embed=embed)
                await makeembed('전송완료!', 'DM전송 기능' + user.name)
                await choose.remove_reaction("⭕",message.author)
                await choose.remove_reaction("⭕",client.user)
                await choose.remove_reaction("❌",client.user)
            if str(reaction.emoji) == '❌':
                await makeembed('메시지', '취소' + user.name)
                await choose.remove_reaction("❌",message.author)
                await choose.remove_reaction("⭕",client.user)
                await choose.remove_reaction("❌",client.user)
                return None

        elif message.content == f'{pre}봇서버':
            now=datetime.datetime.now()
            embed=discord.Embed(
                title=f'봇이 함께하는 서버, 유저 수',
                description=f'서버 : {len(client.guilds)}개\n유저 : {len(client.users)}명'
            )
            embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}건의"):
            now=datetime.datetime.now()
            msg = message.content[4:]
            embed=discord.Embed(
                title=f'{message.guild}에서 {message.author}님의 건의',
                description=f'내용 : {str(msg)}',
                colour=message.author.colour
            )
            embed.set_thumbnail(url=message.author.avatar_url)
            embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
            await client.get_channel(int(723715364634099813)).send(embed=embed)
            await makeembed('건의 전송!', '건의 전송기능')
            if msg == None:
                await makeembed('건의 내용', '건의 내용을 적어주세요!')
                return None

        elif message.content.startswith(f"{pre}Code"):
            msg=message.content.split(' ')
            if msg == None:
                await makeembed('코드입력에 실패했습니다.', '코드를 적어주세요!')
            if msg == 'MHCK728':
                await message.channel.send('WaSans')

        elif message.content.startswith(f'{pre}청소'):
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author.id == 445529063528857611:
                varrr=message.content.split(' ')
                await message.channel.purge(limit=int(varrr[1])+1)
                now=datetime.datetime.now()
                msg=await message.channel.send(embed=discord.Embed(title=f'메시지 {str(int(varrr[1]))}개 삭제 완료!', descirption=f'{client.user.name}이 삭제했어요!!', colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일'))
                await asyncio.sleep(2.5)
                await msg.delete()
            else:
                await message.channel.send(f'{message.author.mention}응 너는 사용불가~')
                return None

        elif message.content.startswith(f'{pre}권한'):
            author = message.mentions[0]
            if author.guild_permissions.administrator: #관리자
                per1 = '관리자 : O'
            else:
                per1 = '관리자 : X'
            
            if author.guild_permissions.view_audit_log: #감사로그 보기
                per2 = '감사로그 보기 : O'
            else:
                per2 = '감사로그 보기 : X'

            if author.guild_permissions.manage_guild: #서버 관리하기
                per3 = '서버 관리하기 : O'
            else:
                per3 = '서버 관리하기 : X'

            if author.guild_permissions.manage_roles: #역할 관리하기
                per4 = '역할 관리하기 : O'
            else:
                per4 = '역할 관리하기 : X'

            if author.guild_permissions.manage_channels: #채널 관리하기
                per5 = '채널 관리하기 : O'
            else:
                per5 = '채널 관리하기 : X'

            if author.guild_permissions.kick_members: #멤버 추방하기
                per6 = '멤버 추방하기 : O'
            else:
                per6 = '멤버 차단하기 : X'

            if author.guild_permissions.ban_members: #멤버 차단하기
                per7 = '멤버 차단하기 : O'
            else:
                per7 = '멤버 차단하기 : X'

            if author.guild_permissions.create_instant_invite: #채널 관리하기
                per8 = '초대 코드 만들기 : O'
            else:
                per8 = '초대 코드 만들기 : X'

            if author.guild_permissions.change_nickname: #별명 변경하기
                per9 = '별명 변경하기 : O'
            else:
                per9 = '별명 변경하기 : X'

            if author.guild_permissions.manage_nicknames: #별명 관리하기
                per10 = '별명 관리하기 : O'
            else:
                per10 = '별명 관리하기 : X'

            if author.guild_permissions.manage_emojis: #이모티콘 관리
                per11 = '이모티콘 관리 : O'
            else:
                per11 = '이모티콘 관리 : X'

            if author.guild_permissions.view_channel: #채팅 채널 읽기 및 음성 채널 보기
                per12 = '채팅 채널 읽기 및 음성 채널 보기 : O'
            else:
                per12 = '채팅 채널 읽기 및 음성 채널 보기 : X'

            if author.guild_permissions.send_messages: #메세지 보내기
                per13 = '메세지 보내기 : O'
            else:
                per13 = '메세지 보내기 : X'

            if author.guild_permissions.send_tts_messages: #tts메세지 보내기
                per14 = 'TTS 메세지 보내기 : O'
            else:
                per14 = 'TTS 메세지 보내기 : X'

            if author.guild_permissions.manage_messages: #메세지 관리하기
                per15 = '메세지 관리하기 : O'
            else:
                per15 = '메세지 관리하기 : X'

            if author.guild_permissions.embed_links: #링크 첨부하기
                per16 = '링크 첨부하기 : O'
            else:
                per16 = '링크 첨부하기 : X'

            if author.guild_permissions.attach_files: #파일 첨부하기
                per17 = '파일 첨부하기 : O'
            else:
                per17 = '파일 첨부하기 : X'

            if author.guild_permissions.read_message_history: #메세지 기록보기
                per18 = '메세지 기록보기 : O'
            else:
                per18 = '메세지 기록보기 : X'

            if author.guild_permissions.mention_everyone: #@everyone, @here 모든 역할 멘션하기
                per19 = '@everyone, @here 모든 역할 멘션하기 : O'
            else:
                per19 = '@everyone, @here 모든 역할 멘션하기 : X'

            if author.guild_permissions.use_external_emojis: #외부 이모티콘 사용하기
                per20 = '외부 이모티콘 사용하기 : O'
            else:
                per20 = '외부 이모티콘 사용하기 : X'

            if author.guild_permissions.add_reactions: #반응 추가하기
                per21 = '반응 추가하기 : O'
            else:
                per21 = '반응 추가하기 : X'

            if author.guild_permissions.connect: #연결
                per22 = '연결 - 음성 : O'
            else:
                per22 = '연결 - 음성 : X'

            if author.guild_permissions.speak: #말하기
                per23 = '말하기 - 음성 : O'
            else:
                per23 = '말하기 - 음성: X'

            if author.guild_permissions.stream: #동영상
                per24 = '동영상 - Go Live : O'
            else:
                per24 = '동영상 - Go Live : X'

            if author.guild_permissions.mute_members: #멤버들의 마이크 음소거하기
                per22 = '멤버들의 마이크 음소거하기 - 음성 : O'
            else:
                per22 = '멤버들의 마이크 음소거하기 - 음성 : X'

            if author.guild_permissions.deafen_members: #멤버의 헤드셋 음소거하기
                per22 = '멤버의 헤드셋 음소거하기 - 음성 : O'
            else:
                per22 = '멤버의 헤드셋 음소거하기 - 음성 : X'

            if author.guild_permissions.move_members: #멤버 이동
                per23 = '멤버 이동 - 음성 : O'
            else:
                per23 = '멤버 이동 - 음성: X'

            if author.guild_permissions.priority_speaker: #말하기
                per23 = '우선 발언권 - 음성 : O'
            else:
                per23 = '우선 발언권 - 음성: X'

            await makeembed(f'{author}님의 {message.guild}에서의 권한', f'{per1}\n{per2}\n{per3}\n{per4}\n{per5}\n{per6}\n{per7}\n{per8}\n{per9}\n{per10}\n{per11}\n{per12}\n{per13}\n{per14}\n{per15}\n{per16}\n{per17}\n{per18}\n{per19}\n{per20}\n{per21}\n{per22}\n{per23}\n{per24}\n')

        elif message.content.startswith(f"{pre}뮤트"):
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author.id == 445529063528857611:
                author = message.mentions[0]
                role = discord.utils.get(message.guild.roles, name="뮤트")
                msg = message.content[26:]
                now=datetime.datetime.now()
                await author.add_roles(role)
                await message.guild.get_channel(message.channel.category_id).set_permissions(author, send_messages=False)
                for channel in message.guild.channels:
                    if (str(channel) in dsv):
                        try:
                            embed = discord.Embed(
                                title=f'{author} is mute from {message.guild} server!',
                                description=f'User : {author} ({author.mention})\nAdmin : {message.author}\nReason : ' + msg,
                                color=discord.Colour.gold()
                            )
                            embed.set_footer(icon_url=author.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                            await channel.send(embed=embed)
                            if not message.embeds == []:
                                await channel.send(embed=message.embeds[0])
                        except:
                            return None
            else:
                await makeembed('사용 ㄴㄴ', '관리자 기능')
                return None

        elif message.content.startswith(f"{pre}언뮤트"):
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author.id == 445529063528857611:
                author = message.mentions[0]
                role = discord.utils.get(message.guild.roles, name="뮤트")
                msg = message.content[27:]
                now=datetime.datetime.now()
                await author.remove_roles(role)
                await message.guild.get_channel(message.channel.category_id).set_permissions(user, overwrite=None)
                for channel in message.guild.channels:
                    if (str(channel) in dsv):
                        try:
                            embed = discord.Embed(
                                title=f'{author} is unmute from {message.guild} server!',
                                description=f'User : {author} ({author.mention})\nAdmin : {message.author}\nReason : ' + msg,
                                color=discord.Colour.green()
                            )
                            embed.set_footer(icon_url=author.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                            await author.remove_roles(role)
                            await channel.send(embed=embed)
                            if not message.embeds == []:
                                await channel.send(embed=message.embeds[0])
                        except:
                            return None
            else:
                await makeembed('사용 ㄴㄴ', '관리자 기능')
                return None

        elif message.content.startswith(f"{pre}경고"):
            if message.author.guild_permissions.administrator or message.author.guild_permissions.manage_messages or message.author.id == 445529063528857611:
                author = message.mention[0]
                wauthor = message.content.split(' ', '<', '>', '@', '!')[0]
                reason = message.content.split(' ')[1]
                file = openpyxl.load_workbook("경고.xlsx")
                sheet = file.active
                i = 1
                while True:
                    if sheet ["A" + str(i)].value == str(author.id):
                        sheet ["B" + str(i)].value = str(sheet ["B" + str(i)].value) + 1
                        file.save("경고.xlsx")
                        if sheet ["B" + str(i)].value == 5:
                            await message.guild.ban(author)
                            embed=discord.Embed(
                                title=f'경고 5회 누적. {wauthor}',
                                description=f'경고 5회 누적으로 서버에서 **추방**됩니다.',
                                colour=discord.Colour.red()
                            ).set_footer(icon_url=wauthor.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                            await message.channel.send(embed=embed)
                        else:
                            embed=discord.Embed(
                                title=f'{wauthor}님의 경고.',
                                description=f'사유 : {reason}\n처리자 : {message.author}\n 처벌 대상 : {wauthor}',
                                colour=discord.Colour.gold()
                            ).set_footer(icon_url=wauthor.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                            await message.channel.send(embed=embed)
                            break
                    if sheet ["A" + str(i)].value == None:
                        sheet ["A" + str(i)].value = str(author.id)
                        sheet ["B" + str(i)].value = 1
                        file.save("경고.xlsx")
                        embed=discord.Embed(
                            title=f'{wauthor}님의 경고.',
                            description=f'사유 : {reason}\n처리자 : {message.author}\n 처벌 대상 : {wauthor}',
                            colour=discord.Colour.gold()
                        ).set_footer(icon_url=wauthor.avatar_url, text=f' | {str(author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                        await message.channel.send(embed=embed)
                        break
                    i += 1
            else:
                await makeembed('사용 ㄴㄴ', '관리자 기능')
                return None

        elif message.content.startswith(f"{pre}캡챠"):
            author = message.author
            role = discord.utils.get(message.guild.roles, name="유저")
            lmage_captcha = ImageCaptcha()
            a = ""
            for i in range(6):
                a += str(random.randint(0, 9))

            name = str(message.author.id) + ".png"
            lmage_captcha.write(a, name)

            await message.channel.send(file=discord.File(name))
            def check(msg):
                return msg.author == message.author and msg.channel == message.channel

            try:
                msg = await client.wait_for("message", timeout=10, check=check)
            except:
                await makeembed('시간초과!', '캡챠인증에 실패했습니다!')
                return

            if msg.content == a:
                await makeembed('성공!', '캡챠인증에 성공했습니다!')
                try:
                    await author.add_roles(role)
                except:
                    return
                await asyncio.sleep(3)
                await message.channel.purge(limit=4)
            else:
                await makeembed('실패!', '캡챠인증에 실패했습니다!')
                await asyncio.sleep(3)
                await message.channel.purge(limit=4)

        elif message.content == f"{pre}핑":
            embed = discord.Embed(color=0x00ff00, title='🏓퐁!')
            embed.add_field(name='Bot server ping : ', value=f' {round(client.latency * 1000)}ms' + f' (0.{round(client.latency * 1000)}seconds)',  inline=True) #inline True : 옆
            embed.set_thumbnail(url=message.author.avatar_url)
            embed.set_footer(text=f"{message.author}", icon_url=message.author.avatar_url)
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}입장"):
            if not message.author.voice:
                await makeembed('사용자는 아직 음성 채널에 들어가있지 않습니다. 들어간 후 시도해주세요.', '오류!')
                return
            channel=message.author.voice.channel
            voice=discord.utils.get(client.voice_clients, guild=message.guild)
            if voice and voice.is_connected():
                await voice.move_to(channel)
            else:
                voice=await channel.connect()
            await voice.disconnect()
            if voice and voice.is_connected():
                await voice.move_to(channel)
            else:
                voice=await channel.connect()
            await makeembed('음성채널 입장 완료', '음성 채널 입장 기능')

        elif message.content==f'{pre}퇴장':
            voice=discord.utils.get(client.voice_clients, guild=message.guild)
            if voice and voice.is_connected():
                await voice.disconnect()
                await makeembed('음성채널 퇴장 완료', '음성 채널 퇴장 기능')
            else:
                await makeembed(f'`{pre}입장`명령어를 사용해야 음악을 들려줄수 있습니다!', '에러')

        elif message.content.startswith(f'{pre}재생'):
            def check_queue():
                Queue_infile=os.path.isdir('./Queue')
                if Queue_infile is True:
                    DIR=os.path.abspath(os.path.realpath('Queue'))
                    length=len(os.listdir(DIR))
                    try:
                        first_file=os.listdir(DIR)[0]
                    except:
                        queues.clear()
                        return
                    main_location=os.path.dirname(os.path.realpath(__file__))
                    song_path=os.path.abspath(os.path.realpath('Queue')+'\\'+first_file)
                    if length!=0:

                        song_there=os.path.isfile('song.mp3')
                        if song_there:
                            os.remove('song.mp3')
                        shutil.move(song_path, main_location)
                        for file in os.listdir('./'):
                            if file.endswith('.mp3'):
                                os.rename(file, 'song.mp3')
                        voice.play(discord.FFmpegPCMAudio('song.mp3'), after=lambda e: check_queue()    )
                        voice.source=discord.PCMVolumeTransformer(voice.source)
                        voice.source.volume=0.20
                    else:
                        queues.clear()
                        return
                else:
                    queues.clear()
            url=message.content[4:]
            song_there=os.path.isfile('song.mp3')
            try:
                if song_there:

                    os.remove('song.mp3')
                    queues.clear()
            except PermissionError:
                await makeembed('음악 재생중입니다', '에러')
                return
            Queue_infile=os.path.isdir('./Queue')
            try:
                Queue_folder='./Queue'
                if Queue_infile is True:
                    shutil.rmtree(Queue_folder)
            except:
                print('')
            await makeembed('노래 다운중...', '잠시만 기다려 주십시오.')
            voice=discord.utils.get(client.voice_clients, guild=message.guild)
            ydl_ops={
                'format':'bestaudio/best',
                'default_search':'ytsearch',
                'postprocessors':[{
                    'key':'FFmpegExtractAudio',
                    'preferredcodec':'mp3',
                    'preferredquality':'192',
                }],
            }
            with youtube_dl.YoutubeDL(ydl_ops) as ydl:
                ydl.download([url])
            for file in os.listdir('./'):
                if file.endswith('.mp3'):
                    name=file
                    os.rename(file, 'song.mp3')
            voice.play(discord.FFmpegPCMAudio('song.mp3'), after=lambda e: check_queue())
            voice.source=discord.PCMVolumeTransformer(voice.source)
            voice.source.volume=1.00
            nname=name.rsplit('-', 2)
            await makeembed(f'{nname} 재생중 <@!{message.author.id}>', '응용 기능')

        elif message.content==f'{pre}일시정지':
            voice=discord.utils.get(client.voice_clients, guild=message.guild)
            if voice and voice.is_playing():
                voice.pause()
                await makeembed(f'일시정지 완료 {message.author.mention}', '응용 기능')
            else:
                await makeembed('음악을 재생하지 않았습니다.', '에러')

        elif message.content==f'{pre}다시재생':
            voice=discord.utils.get(client.voice_clients, guild=message.guild)
            if voice and voice.is_paused():
                await makeembed(f'다시재생 완료 {message.author.mention}', '응용 기능')
                voice.resume()
            else:
                await makeembed('`!일시정지`를 하고 시도해 주십시오.', '에러')

        elif message.content==f'{pre}스킵':
            voice=discord.utils.get(client.voice_clients, guild=message.guild)
            if voice and voice.is_playing():
                queues.clear()
                await makeembed(f'스킵 완료 {message.author.mention}', '응용 기능')
                voice.stop()
            else:
                await makeembed('음악을 재생중이지 않습니다.', '에러 감지 기능')

        elif message.content.startswith(f'{pre}다음'):
            url=message.content[7:]
            Queue_infile=os.path.isdir("./Queue")
            if Queue_infile is False:
                os.mkdir("Queue")
            DIR=os.path.abspath(os.path.realpath("Queue"))
            q_num=len(os.listdir(DIR))
            q_num+=1
            add_queue=True
            while add_queue:
                if q_num in queues:
                    q_num+=1
                else:
                    add_queue=False
                    queues[q_num]=q_num
            queue_path=os.path.abspath(os.path.realpath("Queue")+f'/song{q_num}.%(ext)s')
            ydl_ops={
                'format':'bestaudio/best',
                'default_search':'ytsearch',
                'quite':True,
                'outtmpl':queue_path,
                'postprocessors':[{
                    'key':'FFmpegExtractAudio',
                    'preferredcodec':'mp3',
                    'preferredquality':'192',
                }],
            }
            with youtube_dl.YoutubeDL(ydl_ops) as ydl:
                ydl.download([url])
            await makeembed(f'재생목록에 음악이 추가되었습니다 {message.author.mention}', f'{client.user.name} 노래 기능')

        elif message.content.startswith(f'{pre}공지'):
            if message.author.id!=445529063528857611:
                await makeembed('사용 ㄴㄴ', '관리자 기능')
                return None
            msg=message.content[4:]
            now=datetime.datetime.now()
            embed=discord.Embed(
                title=msg.split('￣')[0],
                description=msg.split('￣')[1] + f'\n\n:link:[스카이봇 포럼](https://discord.gg/57cSTKB)\n:link:[코어 엔터테인먼트](https://discord.gg/TeCpcBq)\n:link:[{client.user.name} 초대](http://bitly.kr/Jsc357T0v)\n:link:[{client.user.name} 하트누르기](https://koreanbots.dev/bots/689419260866330645)',
                colour=discord.Colour.green()
            ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
            embed.set_thumbnail(url=message.author.avatar_url)
            for i in client.guilds:
                arr=[0]
                alla=False
                flag=True
                z=0
                for j in i.channels:
                    arr.append(j.id)
                    z+=1
                    if '📢봇-공지' in j.name or '봇-공지' in j.name or '📢봇-공지사항' in j.name or '아파트-봇-공지사항' in j.name or '📢봇_공지' in j.name or '📢봇공지' in j.name or 'notice' in j.name or '📢『공지ㆍ봇_공지』' in j.name or '봇-공지📜' in j.name:
                        if str(j.type)=='text':
                            try:
                                await j.send(embed=embed)
                                alla=True
                            except:
                                pass
                            break
                if alla==False:
                    try:
                        chan=i.channels[1]
                    except:
                        pass
                    if str(chan.type)=='text':
                        try:
                            await chan.send(embed=embed)
                        except:
                            pass
            await makeembed('공지 전송을 완료했습니다', '관리자 기능')
            print(f'{flag}')

        elif message.content == f'{pre}시간':
            now=datetime.datetime.now()
            embed=discord.Embed(
                title=f'현재 시간',
                descroption=f'{str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초',
                colour=discord.Colour.blue()
            ).set_footer(icon_url=message.author.avatar_url, text=f'{str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
            await message.channel.send(embed=embed)

        elif message.content == (f"{pre}명령어"):
            embed=discord.Embed(colour=0x00ff00, timestamp=message.created_at, title=f'{client.user.name}명령어')
            embed.add_field(name=f'`{pre}정보`',  value=f'사용법 : {pre}정보 @~~~', inline=False)
            embed.add_field(name=f'`{pre}DM`', value=f'사용법 : {pre}DM <멘션> <내용>', inline=False)
            embed.add_field(name=f'`{pre}핑`', value=f'사용법 : {pre}핑', inline=False)
            embed.add_field(name=f'`{pre}캡챠`', value=f'사용법 : {pre}캡챠', inline=False)
            embed.add_field(name=f'`{pre}마크서버`', value=f'사용법 : {pre}마크서버 <서버주소>', inline=False)
            embed.add_field(name=f'`{pre}음악`', value=f'사용법 : {pre}음악')
            embed.add_field(name=f'`{pre}건의`', value=f'사용법 : {pre}건의 <내용>', inline=False)
            embed.add_field(name=f'`{pre}프사`', value=f'사용법 : {pre}프사 <멘션>', inline=False)
            embed.add_field(name=f'`{pre}서버정보`', value=f'사용법 : {pre}서버정보')
            embed.add_field(name=f'`{pre}역할`', value=f'사용법 : {pre}역할 <역할이름> <멘션>')
            embed.add_field(name=f'`{pre}시간`', value=f'사용법 : {pre}시간')
            embed.add_field(name=f'`{pre}실검`', value=f'사용법 : {pre}실검')
            embed.add_field(name=f'`{pre}가위바위보`', value=f'사용법 : {pre}가위바위보 (가위, 바위, 보 중 하나)')
            embed.add_field(name=f'`{pre}채널만들기/{pre}보이스챗`', value=f'사용법 : {pre}채널만들기/{pre}보이스챗 <이름>')
            embed.add_field(name=f'`{pre}건의`', value=f'사용법 : {pre}건의 <내용>')
            embed.add_field(name=f'`{pre}오늘뭐먹지`', value=f'사용법 : {pre}오늘뭐먹지')
            embed.add_field(name=f'`{pre}명령어 관리진`', value=f'사용법 : {pre}명령어 관리진', inline=False)
            embed.add_field(name=f'???', value='\n\n\n[스카이봇 포럼](https://discord.gg/57cSTKB)\n[코어 엔터테인먼트](https://discord.gg/TeCpcBq)', inline=False)
            embed.set_thumbnail(url=message.author.avatar_url)
            embed.set_footer(text=f"{message.author}", icon_url=message.author.avatar_url)
            await message.channel.send(embed=embed)

        elif message.content == f"{pre}명령어 관리진":
            now=datetime.datetime.now()
            embed=discord.Embed(colour=0x00f000, timestamp=message.created_at, title='관리진 전용 명령어')
            embed.add_field(name=f'`{pre}공지`',  value=f'사용법 : {pre}공지 내용/내용', inline=True)
            embed.add_field(name=f'`{pre}채널메시지`', value=f'사용법 : {pre}채널메시지 <채널멘션> (내용)', inline=True)
            embed.add_field(name=f'`{pre}경고`', value=f'사용법 : {pre}경고 <ID>')
            embed.add_field(name=f'`{pre}뮤트/~언뮤트`', value=f'사용법 : {pre}뮤트/{pre}언뮤트 <멘션>')
            embed.add_field(name=f'`{pre}청소`', value=f'사용법 : {pre}청소 <숫자> (두번 뛰어야 합니다.)')
            embed.add_field(name=f'`{pre}ban/{pre}kick`', value=f'사용법 : {pre}ban/{pre}kick <멘션>')
            embed.add_field(name=f'`{pre}개별공지`', value=f'사용법 : {pre}개별공지 <멘션> (악성 문자를 보내지 마시오.)')
            embed.set_thumbnail(url=message.author.avatar_url)
            embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
            await message.channel.send(embed=embed)

        elif message.content == f"{pre}숨겨진명령어":
            embed = discord.Embed(color=0x00ff00)
            embed.add_field(name=f'{client.user.name}!', value=f'{message.author}', inline=True)
            embed.set_footer(text=f'{message.author}', icon_url=message.author.avatar_url)
            await message.channel.send(embed=embed)
                                
        elif message.content == f"{pre}와아아아앗더":
            await makeembed('와아아아', '앗더')

        elif message.content == f"{pre}음악":
            embed=discord.Embed(colour=0x00f000, timestamp=message.created_at)
            embed.add_field(name=f'`{pre}입장`',  value='자신이 있는 음성채널에 입장합니다.', inline=True)
            embed.add_field(name=f'`{pre}퇴장`', value='음성채널에서 퇴장합니다.', inline=True)
            embed.add_field(name=f'`{pre}재생`', value='음악을 재생합니다.', inline=True)
            embed.add_field(name=f'`{pre}일시정지`', value='음악을 일시정지 합니다', inline=True)
            embed.add_field(name=f'`{pre}다시재생`', value='일시정지했던 음악을 다시 재생합니다', inline=True)
            embed.add_field(name=f'`{pre}스킵`', value='틀고있던 노래를 스킵하고 다음노래로 이동합니다(?)', inline=True)
            embed.add_field(name=f'`{pre}다음`', value='다음 재생목록에 추가되있던 노래를 틉니다.')
            embed.set_footer(text=f"{message.author}", icon_url=message.author.avatar_url) #스킵, 다음
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}이스터에그"):
            embed=discord.Embed(colour=0x00f000, timestamp=message.created_at)
            embed.add_field(name=f'`{pre}이스터에그`',  value='이스터에스 명령어를 확인합니다..', inline=True)
            embed.add_field(name=f'`{pre}', value='?', inline=True)
            embed.add_field(name=f'`{pre}숨겨진명령어`', value='?', inline=True)
            embed.add_field(name=f'`{pre}discord`', value='?', inline=True)
            embed.add_field(name=f'`{pre}SCP`', value='?', inline=True)
            embed.set_thumbnail(url=message.author.avatar_url)
            embed.set_footer(text=f"{message.author}", icon_url=message.author.avatar_url)
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}discord"):
            await makeembed('이제...', '`Skype와 TeamSpeak`와 이별할 시간이예요!')
            await makeembed('디스코드라...', '(욕하면 죽을듯...)')

        elif message.content.startswith(f"{pre}고급설정"):
            msg = message.content[6:]
            if msg == '도배':
                await makeembed(f'{msg}', '도배기능')
                i = 0
                while True:
                    i += 1
                    await message.author.send('도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배도배')
                    if i > 200:
                        break
                await message.author.send(f'|| 어때요? 저의 도배실력이? <@{message.author.id}> ||')
                await makeembed(f'어때요? 저의 도배실력이? {message.author.mention}', '|| 하하하하! ||')
            if msg == '인사':
                await makeembed(f'인사 좋아해요? {message.author.mention}', '|| 안녕하세요! 저는 <@!689419260866330645>에요! ||')
            if msg == '소개':
                await makeembed(f'안녕하세욧!', f'전 {client.user.name}이에욧!')

        elif message.content.startswith(f"{pre}1238dsfkdksksl293도djfds배sdf"):
            author = message.mentions[0]
            msg=message.content.split(' ')[1]
            if message.author.id == 524515155254444032 or 445529063528857611:
                await message.delete()
                i = 0
                while True:
                    i += 1
                    await author.send(f'{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}{msg}')
                    if i > 400:
                        break
                await message.author.send(f' {message.author.mention} 님! 제가 {author} 님한테 도배를 완료했어요!')
            else:
                await makeembed('알수없는 명령어', '알수없는 명령어입니다.')
                return None
        
        elif message.content.startswith(f"{pre}???"):
            await makeembed('???', '???')

        elif message.content == f'{pre}종료':
            if message.author.id==445529063528857611:
                now=datetime.datetime.now()
                choose=await makeembed('종료',f'{client.user.name}을 진짜 종료하시겠어요?')
                await choose.add_reaction("⭕")
                await choose.add_reaction("❌")
                def c(reaction, user):
                    return user == message.author and str(reaction) in ["⭕", "❌"] and choose.id == reaction.message.id # 이모지 리액션 부분
                reaction, user = await client.wait_for('reaction_add',check=c)# 이모지 리액션 부분
                if str(reaction.emoji) == '⭕':
                    embed=discord.Embed(
                        title=f'❌종료❌',
                        description=f'❌_❌ {client.user.name}을 종료합니다.',
                        colour=discord.Colour.red()
                    )
                    embed.set_footer(icon_url=client.user.avatar_url, text=f'{str(now.year)}년 {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초')
                    await client.get_channel(int(727393651357122562)).send(embed=embed)
                    await choose.remove_reaction("⭕",message.author)
                    await choose.remove_reaction("⭕",client.user)
                    await choose.remove_reaction("❌",client.user)
                    await asyncio.sleep(2.5)
                    await client.logout()
                if str(reaction.emoji) == '❌':
                    await makeembed('종료', '취소 : ' + user.name)
                    await choose.remove_reaction("❌",message.author)
                    await choose.remove_reaction("⭕",client.user)
                    await choose.remove_reaction("❌",client.user)
                    return None
            else:
                await makeembed(f'{author}님...', '어디서 저를 종료하려고 하세요... 에휴...')
                return None

        elif message.content == f'{pre}사진':
            embed=discord.Embed(colour=message.author.colour, timestamp=message.created_at)
            embed.add_field(name='명령어 : ', value='명령어 입니다.', inline=True)
            embed.set_image(url='http://bitly.kr/NA7Xz72z')
            embed.set_footer(text=f'{message.author}', icon_url=message.author.avatar_url)
            await message.channel.send(embed=embed)

        elif message.content.startswith(f'{pre}Qr'):
            target=message.content[4:]
            client_id="TmlZec6PQAZrzVq_fIao"
            client_secret="fzsgH99SP0"
            header = {'X-Naver-Client-Id': client_id, 'X-Naver-Client-Secret': client_secret}
            naver = 'https://openapi.naver.com/v1/util/shorturl'
            data = {'url': target}
            maker=requests.post(url=naver,data=data,headers=header)
            maker.close()
            output=maker.json()['result']['url']
            qr=qrcode.QRCode(
                version=1,
                box_size=13,
                border=2

            )
            data=output
            qr.add_data(data)
            qr.make(fit=True)
            img=qr.make_image(fill="black", back_color="white")
            img.save(f'{message.author.id}' + '.png')
            name = str(f'{message.author.id}' + '.png')
            embed=discord.Embed(title=f'{message.author}님이 요청하신 QR코드입니다.\n요청하신 URL : {output}\nURL이 길수도 있어서 단축URL로 QR코드를 생성합니다', description='QR코드를 우클릭해서 이미지저장하세요!!', colour=message.author.colour, timestamp=message.created_at)
            embed.set_footer(text=f'{message.author}', icon_url=message.author.avatar_url)
            await message.channel.send(embed=embed)
            await message.channel.send(file=discord.File(name))

        elif message.content.startswith(f'{pre}마크서버'):
            server=message.content.split(' ')[1]
            url='https://mcsrvstat.us/server/'+str(server)
            res=requests.get(   url)
            soup=BeautifulSoup(res.content, 'html.parser')
            clas=soup.findAll('h2')
            try:
                adsf=clas[1]
                if adsf==None:
                    return
            except IndexError:
                await makeembed('✅ 서버가 온라인입니다. ✅', '✅ 마인크래프트 서버 온/오프라인 확인기능 ⛔')
                return
            await makeembed(f'⛔ {str(server)} 서버가 오프라인입니다. ⛔', f'✅ {str(server)} 마인크래프트 서버 온/오프라인 확인기능 ⛔')

        elif message.content == f'{pre}봇정보':
            now=datetime.datetime.now()
            embed=discord.Embed(
                title=f'{client.user.name} 정보',
                description=f'> 봇 제작 : UN10#7291 (Team SB소속)\n> 도움을 주신분 : <@!524515155254444032>\n> 봇 id : 689419260866330645\n> 봇 생일 : 2020년 03월 17일\n> :link:[초대](http://bitly.kr/Jsc357T0v)\n \n> 서버 수 : {len(client.guilds)}개\n> 유저 수 : {len(client.users)}명\n \n> Python : Python 3.8.5 32-bit\n> discord.py : discord.py 1.3.4',
                colour=discord.Colour.green()
                )
            embed.set_thumbnail(url=client.user.avatar_url)
            embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
            await message.channel.send(embed=embed)
            embed=discord.Embed(
                title=f'{client.user.name}에게 힘 주기',
                description=f'[:heart: 주로가기](https://koreanbots.dev/bots/689419260866330645)',
                color=0xFFA9
            )
            await message.channel.send(embed=embed)

        elif message.content.startswith(f"{pre}영한번역"):
                channel = message.channel
                url="https://openapi.naver.com/v1/papago/n2mt?source=en&target=ko&text="
                text = message.content[6:]
                if text == "":
                    await makeembed('번역할 내용를 입력해주세요', '오류!')
                request_url = "https://openapi.naver.com/v1/papago/n2mt"
                headers= {"X-Naver-Client-Id": "TmlZec6PQAZrzVq_fIao", "X-Naver-Client-Secret":"fzsgH99SP0"}
                params = {"source": "en", "target": "ko", "text": text}
                response = requests.post(request_url, headers=headers, data=params)
                result = response.json()
                result = result['message']['result']['translatedText']
                await makeembed(':flag_us::flag_kr:번역된 문장',f':flag_us:영어```{text}```:flag_kr:한국어\n```{result}```')
    
        elif message.content.startswith(f"{pre}영일번역"):
            channel = message.channel
            url="https://openapi.naver.com/v1/papago/n2mt?source=en&target=ja&text="
            text = message.content[6:]
            if text == "":
                await makeembed('번역할 내용를 입력해주세요', '오류!')
            request_url = "https://openapi.naver.com/v1/papago/n2mt"
            headers= {"X-Naver-Client-Id": "TmlZec6PQAZrzVq_fIao", "X-Naver-Client-Secret":"fzsgH99SP0"}
            params = {"source": "en", "target": "ja", "text": text}
            response = requests.post(request_url, headers=headers, data=params)
            result = response.json()
            result = result['message']['result']['translatedText']
            await makeembed(':flag_us::flag_jp:번역된 문장',f':flag_us:영어```{text}```:flag_jp:일본어\n```{result}```')

        elif message.content.startswith(f"{pre}영중번역"):
            channel = message.channel
            url="https://openapi.naver.com/v1/papago/n2mt?source=en&target=zh-CN&text="
            text = message.content[6:]
            if text == "":
                await makeembed('번역할 내용를 입력해주세요', '오류!')
            request_url = "https://openapi.naver.com/v1/papago/n2mt"
            headers= {"X-Naver-Client-Id": "TmlZec6PQAZrzVq_fIao", "X-Naver-Client-Secret":"fzsgH99SP0"}
            params = {"source": "en", "target": "zh-CN", "text": text}
            response = requests.post(request_url, headers=headers, data=params)
            result = response.json()
            result = result['message']['result']['translatedText']
            await makeembed(':flag_us::flag_cn:번역된 문장',f':flag_us:영어```{text}```:flag_cn:중국어\n```{result}```')

        elif message.content.startswith(f"{pre}한영번역"):
            channel = message.channel
            url="https://openapi.naver.com/v1/papago/n2mt?source=ko&target=en&text="
            text = message.content[6:]
            if text == "":
                await makeembed('번역할 내용를 입력해주세요', '오류!')
            request_url = "https://openapi.naver.com/v1/papago/n2mt"
            headers= {"X-Naver-Client-Id": "TmlZec6PQAZrzVq_fIao", "X-Naver-Client-Secret":"fzsgH99SP0"}
            params = {"source": "ko", "target": "en", "text": text}
            response = requests.post(request_url, headers=headers, data=params)
            result = response.json()
            result = result['message']['result']['translatedText']
            await makeembed(':flag_kr::flag_us:번역된 문장',f':flag_kr:한국어```{text}```:flag_us:영어\n```{result}```')

        elif message.content.startswith(f"{pre}한일번역"):
            channel = message.channel
            url="https://openapi.naver.com/v1/papago/n2mt?source=ko&target=ja&text="
            text = message.content[6:]
            if text == "":
                await makeembed('번역할 내용를 입력해주세요', '오류!')
            request_url = "https://openapi.naver.com/v1/papago/n2mt"
            headers= {"X-Naver-Client-Id": "TmlZec6PQAZrzVq_fIao", "X-Naver-Client-Secret":"fzsgH99SP0"}
            params = {"source": "ko", "target": "ja", "text": text}
            response = requests.post(request_url, headers=headers, data=params)
            result = response.json()
            result = result['message']['result']['translatedText']
            await makeembed(':flag_kr::flag_jp:번역된 문장',f':flag_kr:한국어```{text}```:flag_jp:일본어\n```{result}```')

        elif message.content.startswith(f"{pre}한중번역"):
            channel = message.channel
            url="https://openapi.naver.com/v1/papago/n2mt?source=ko&target=zh-CN&text="
            text = message.content[6:]
            if text == "":
                await makeembed('번역할 내용를 입력해주세요', '오류!')
            request_url = "https://openapi.naver.com/v1/papago/n2mt"
            headers= {"X-Naver-Client-Id": "TmlZec6PQAZrzVq_fIao", "X-Naver-Client-Secret":"fzsgH99SP0"}
            params = {"source": "ko", "target": "zh-CN", "text": text}
            response = requests.post(request_url, headers=headers, data=params)
            result = response.json()
            result = result['message']['result']['translatedText']
            await makeembed(':flag_kr::flag_cn:번역된 문장',f':flag_kr:한국어```{text}```:flag_cn:중국어\n```{result}```')
    except ZeroDivisionError:
        None
    except discord.errors.Forbidden:
        None
    except IndexError:
        await makeembed('입력형식을 맞게 적어주세요!', '에러 감지 기능')
    except SyntaxError:
        await makeembed('오타를 확인해주세요!', '에러 감지 기능')
    except Exception as ex:
        embed=discord.Embed(
            title=f':x: 오류! **{message.content}** :x:',
            description=f'{message.guild} ({message.guild.id})\n\nUser : {message.author.mention} ({message.author.id})\n\nchannel : {message.channel.mention} ({message.channel.id})\n\nError : {ex}'
        )
        embed.set_footer(text=f'{client.user.name} | {str(now.month)}월 {str(now.day)}일 {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초', icon_url=message.author.avatar_url)
        await message.channel.send(embed=embed)
        await client.get_channel(int(727403360101203970)).send(embed=embed)
        await client.get_channel(int(727403360101203970)).send(f'```py\n{ex}```')
async def my_background_task():
    await client.wait_until_ready()
    while not client.is_closed():
        act=["=명령어로 명령어 확인", f'{len(client.guilds)}개의 서버에 참여', f'{len(client.users)}명의 유저들과 소통', 'Visual Studio Code ', '문의 : UN10#7291', '팀 SB', 'python 3.8.3', f'나는야 {client.user.name}', '인증하고 싶다', '빨리 75개 좀 되라ㅏㅏㅏㅏ']
        for i in act:
            game = discord.Game(str(i))
            await client.change_presence(status=discord.Status.dnd, activity=game)
            await asyncio.sleep(5)
client.loop.create_task(my_background_task())
client.run(token)
os.execv(sys.executable, ['python']+sys.argv)
